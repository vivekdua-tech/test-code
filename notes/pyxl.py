import openpyxl as xl
import os


os.chdir('/Users/vidua/pyclass/notes')
wb = xl.load_workbook('Book2.xlsx')
sheet = wb['Sheet1']

prices = []

for row in range(4, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    prices.append(cell.value)

dis_prices = [0.90 * x for x in prices]

for index in range(len(dis_prices)):
    cell = sheet.cell(4 + index, 4)
    cell.value = dis_prices[index]

wb.save('Book3.xlsx')
